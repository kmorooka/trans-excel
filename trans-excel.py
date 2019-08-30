#!/usr/bin/python 
# coding:utf-8
"""
File: trans.py
Fuction: Translate English documents to Japanese.
Version: 0.2 (As of Aug.2019)
"""
import sys
import shutil
import boto3
import openpyxl
from openpyxl import load_workbook

SOURCE_LANG = "en"
TARGET_LANG = "ja"
translate = boto3.client(service_name='translate')
# ----------------------------------------------------
# cp_file(arg1, arg2)
#     arg1: Original English Excel file name.
#     arg2: Translate target Japanese Excel file name.
#     return: none.
# ----------------------------------------------------
def cp_file(fn_in, fn_out):
    shutil.copy(fn_in, fn_out)
    return()
    
# ----------------------------------------------------
# trans_excel(arg)
#    arg: input sentence for translate. Maybe English
#    return: translated sentence. Maybe Japanese
# ----------------------------------------------------
def trans_excel(phrase):
    result = translate.translate_text(Text=str(phrase), SourceLanguageCode=SOURCE_LANG, TargetLanguageCode=TARGET_LANG)
    return(result.get('TranslatedText'))
    
# ----------------------------------------------------
# proc_excel(arg1)
#    arg1: File name for writing traslated sentence.
# ----------------------------------------------------
def proc_excel(fn):
    wb = load_workbook(fn)  #--- ALl sheet data are stored to wb.
    
    for ws in wb:  #---Process Each sheet.
        print('Processing - Sheet Title = %s' % ws.title)
        for rr in range(1, ws.max_row):  #Read each cell > Translate > Replace cell in JA.
            for cc in range(1, ws.max_column):
                tmp = ws.cell(column=cc, row=rr).value  #Read Cell value.
                if tmp is None:  #Ignore blank cell.
                    continue
                else:
                    ja_str = trans_excel(tmp)  #Translate en to ja.
                    ws.cell(column=cc, row=rr, value=ja_str)  #Replace from en to ja.
    wb.save(fn)
    wb.close()
    return()
# ----------------------------------------------------
# main() 
#    arg1: Original Excel file name.
#    arg2: Target Excel file name. File created if there not exist. Overwrite if there is exist.
# ----------------------------------------------------
def main():
    
    args = sys.argv
    num = len(args)
    if num != 3:
        print('trans: ERROR: Usage $./trans.py <Input Excel name> <Output Excel name>')
        sys.exit('trans: ERROR.')
    rc = cp_file(args[1], args[2])  # cp & open output file.
    proc_excel(args[2])
    
    sys.exit('trans: END Successfully.')

# ----------------------------------------------------
if __name__ == '__main__':
    main()
    
# ----------------------------------------------------
# End of File
# ----------------------------------------------------
